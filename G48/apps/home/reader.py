# -*- coding: utf-8 -*-
import os
import xlrd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from zipfile import BadZipfile
import unicodecsv as csv
import chardet
import time


def toUnicode(x):
    if not isinstance(x, unicode):
        x = unicode(x)
    return x


def xlsxToUnicode(x):
    y = x.value
    if not isinstance(y, unicode):
        y = unicode(y)
    return y


def read_file_xls(file_name):
    '''
    读取xls表格内容，按照格式返回信息
    '''
    if not os.path.isfile(file_name):
        print '%s不是文件', file_name

    real_file_name = os.path.basename(file_name)       # 得到一个路径下的文件名
    res = {}
    res['tname'] = file_name                           # 存储表名
    res["bad_file_info"] = {}  # 如果是一个坏文件，则存储
    try:
        workbook = xlrd.open_workbook(file_name)
    except BadZipfile:
        print "坏文件："
        print file_name
        res["bad_file_info"] = {"error_type": "BadZipfile", "file_name": file_name}
        return res
    except TypeError:
        print "类型出错文件："
        print file_name
        res["bad_file_info"] = {"error_type": "TypeError", "file_name": file_name}
        return res
    except IOError:
        print "文件正在被其他应用打开: "
        print file_name
        res["bad_file_info"] = {"error_type": "IOError", "file_name": file_name}
        return res
    except:
        print "其它错误："
        res["bad_file_info"] = {"error_type": "OtherError", "file_name": file_name}
        return res

    sheets = workbook.sheets()

    xlsData = {}
    for sheet in sheets:
        sheet_data = {}
        rows = sheet.nrows
        cols = sheet.ncols

        # 有些sheet是空的，防止这种现象
        if rows == 0:
            continue
        sheet_data['sname'] = sheet.name                # 存储sheet名
        sheet_data['content'] = [map(toUnicode, sheet.row_values(row)) for row in range(0, rows)]  # 从第一行开始存储

        # 解析表头数据,表头一般在前三行中
        # 统计前三行中每行非空元素的个数,个数最多的定义为表头(这个寻找表头的方式可进一行探讨).
        head_len_list = []
        row_nums = len(sheet_data['content'])  # 总行数
        head_nums = 0  # 表头可能在前head_nums行
        if row_nums > 3:
            head_nums = 3
        else:
            head_nums = row_nums
        for i in range(head_nums):
            length = len(sheet_data['content'][i])
            for x in sheet_data['content'][i]:
                if not x or x in [None, u'None']:
                    length -= 1
            head_len_list.append(length)
        head_index = head_len_list.index(max(head_len_list))
        sheet_data['header'] = sheet_data['content'][head_index]

        # sheet_data['header'] = sheet_data['content'][0]
        sheet_data['rows'] = rows
        sheet_data['cols'] = cols
        xlsData[sheet.name] = sheet_data                # 存储数据内容
    res['sheets'] = xlsData                             # 存储所有的sheet
    return res


def read_file_xlsx(file_name):
    '''
    读取xls表格内容，按照格式返回信息
    '''
    if not os.path.isfile(file_name):
        print '%s不是文件', file_name

    real_file_name = os.path.basename(file_name)       # 得到一个路径下的文件名
    res = {}
    res["bad_file_info"] = {}                          # 如果是一个坏文件，则存储
    res['tname'] = file_name                           # 存储表名
    sheets = []
    try:
        # start = time.clock()
        workbook = load_workbook(file_name)            # 得到一个excel表的对象
        # end = time.clock()
        # if end - start > 4:
        #     print "4s:文件:", file_name, "时间：", end-start
        #     print "**************************"
        sheets = workbook.worksheets                   # 所有的sheet表
    except BadZipfile:
        print "坏文件："
        print file_name
        res["bad_file_info"] = {"error_type": "BadZipfile", "file_name": file_name}
        return res
    except TypeError:
        print "类型出错文件："
        print file_name
        res["bad_file_info"] = {"error_type": "TypeError", "file_name": file_name}
        return res
    except IOError:
        print "文件正在被其他应用打开: "
        print file_name
        res["bad_file_info"] = {"error_type": "IOError", "file_name": file_name}
        return res
    except:
        print "其它错误："
        res["bad_file_info"] = {"error_type": "OtherError", "file_name": file_name}
        return res

    xlsData = {}
    for sheet in sheets:
        sheet_data = {}
        rows = sheet.max_row
        cols = sheet.max_column
        if rows == 0:
            continue
        sheet_data['sname'] = sheet.title                                          # 存储sheet名
        sheet_data['content'] = [map(xlsxToUnicode, line) for line in list(sheet.rows)]  # 从第一行开始存储

        # 解析表头数据,表头一般在前三行中
        # 统计前三行中每行非空元素的个数,个数最多的定义为表头(这个寻找表头的方式可进一行探讨).
        head_len_list = []
        row_nums = len(sheet_data['content'])  # 总行数
        head_nums = 0                          # 表头可能在前head_nums行
        if row_nums > 3:
            head_nums = 3
        else:
            head_nums = row_nums
        for i in range(head_nums):
            length = len(sheet_data['content'][i])
            for x in sheet_data['content'][i]:
                if not x or x in [None, u'None']:
                    length -= 1
            head_len_list.append(length)
        head_index = head_len_list.index(max(head_len_list))
        sheet_data['header'] = sheet_data['content'][head_index]

        sheet_data['rows'] = rows
        sheet_data['cols'] = cols
        xlsData[sheet.title] = sheet_data                # 存储数据内容
    res['sheets'] = xlsData                             # 存储所有的sheet
    return res


def read_file_csv(file_name):
    '''
    读取xls表格内容，按照格式返回信息
    '''
    if not os.path.isfile(file_name):
        print '%s不是文件' % file_name

    real_file_name = os.path.basename(file_name)       # 得到一个路径下的文件名
    # 获取文件的编码方式
    csvfile_test_code = open(file_name, 'r')
    file_encoding = chardet.detect(csvfile_test_code.read())  # 获取文件的编码方式
    csvfile_test_code.close()

    # 将GB2312编码改成GBK，以免有中文扩展集导致编码错误
    if file_encoding['encoding'] == "GB2312":
        file_encoding['encoding'] = "GBK"

    res = {}
    xlsData = {}
    res['tname'] = file_name                            # 存储表名
    res["bad_file_info"] = {}                           # 如果是一个坏文件，则存储

    try:
        with open(file_name, 'r') as csvfile:
            try:
                reader = csv.reader(csvfile, encoding=file_encoding['encoding'])
                sheet_data = {}
                rows = 0
                cols = 0
                raw_data = [row for row in reader]

                # 解析表头数据,表头一般在前三行中
                # 统计前三行中每行非空元素的个数,个数最多的定义为表头(这个寻找表头的方式可进一行探讨).
                row_nums = len(raw_data)  # 总行数
                head_nums = 0  # 表头可能在前head_nums行
                if row_nums > 3:
                    head_nums = 3
                else:
                    head_nums = row_nums
                head_len_list = []
                for i in range(head_nums):
                    length = len(raw_data[i])
                    for x in raw_data[i]:
                        if not x or x in [None, u'None']:
                            length -= 1
                    head_len_list.append(length)
                head_index = head_len_list.index(max(head_len_list))
                head_row = raw_data[head_index]

                for row in raw_data:                                # 从第一行开始遍历
                    rows += 1
                    if cols < len(row):
                        cols = len(row)
                sheet_data['sname'] = os.path.splitext(real_file_name)[0]   # 存储sheet名
                sheet_data['header'] = head_row       # 存储表头

                content = []
                for line in raw_data:
                    line = [s.strip() for s in line]
                    m, n = len(line), cols
                    if m < n:
                        line.extend(["" for i in range(n - m)])
                    content.append(line)
                sheet_data["content"] = content
                sheet_data['rows'] = rows
                sheet_data['cols'] = cols
                xlsData[sheet_data['sname']] = sheet_data  # 存储数据内容
            # except UnicodeDecodeError:
            #     print '编码出错的文件：'
            #     print file_name
            #     return res
            except BadZipfile:
                print "坏文件："
                print file_name
                res["bad_file_info"] = {"error_type": "BadZipfile", "file_name": file_name}
                return res
            except TypeError:
                print "类型出错文件："
                print file_name
                res["bad_file_info"] = {"error_type": "TypeError", "file_name": file_name}
                return res
            except:
                res["bad_file_info"] = {"error_type": "OtherError", "file_name": file_name}
                return res
    except IOError:
        print "文件正在被其他应用打开: "
        print file_name
        res["bad_file_info"] = {"error_type": "IOError", "file_name": file_name}
        return res
    except:
        print "其它错误文件："
        res["bad_file_info"] = {"error_type": "OtherError", "file_name": file_name}
        return res
    res['sheets'] = xlsData  # 存储所有的sheet
    return res
